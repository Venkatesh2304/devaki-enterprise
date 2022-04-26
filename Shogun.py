
import pandas as pd
from selenium import webdriver
import os
import datetime
import time
from selenium.webdriver.common.by import By
import win32com.client
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from openpyxl  import load_workbook
from win32api import keybd_event
import re
import sys
import shutil
        
#import win32com.client as win32

def dispatch(app_name:str):
    try:
        app = win32com.client.gencache.EnsureDispatch(app_name)
    except AttributeError:
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
      
        app = win32com.client.gencache.EnsureDispatch(app_name)
    return app

def waitlog(reason,timeout=30) :
  global driver
  for i in range(timeout) :
   x=driver.execute_script('return document.querySelector("#display-message").innerText')
   if reason.lower() in x.lower() :
       time.sleep(1)
   else :
       break
  if i==30 :
    try :
     driver.execute_script('document.querySelector("body > div.panel.window.ui-draggable.ui-resizable.ui-resizable-disabled.messager-window > div.dialog-button.messager-button > a").click();')    
     time.sleep(1)
     driver.execute_script('return document.querySelector("#display-message").innerText')
    except :
     x=2/0
     
def waituntil(x) :
 global driver
 for i in range(60):
  try :
    print(x)
    return driver.execute_script(x)
    break
  except:
    time.sleep(1)
    print(1)
 if i==60 :
   x=1/0
def automate() :
    path='D:\\basepack\\'
    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--start-maximized")
    prefs = {'download.default_directory' : path }
    options.add_experimental_option('prefs', prefs)
    global driver
    driver = webdriver.Chrome(r'chromedriver.exe',options=options)
    return automates(driver)
    
def reader1(x,y) :
    global df
    df=pd.read_excel(x+y)
def reader2(x,y) :
    global df_2
    df_2=pd.read_excel(x+y,usecols=['Unnamed: 1','Bill Wise Sales'])

def automates(driver) :
 path='D:\\basepack\\'
 f=open('login.txt')
 userdata=eval(f.read())
 f.close()
 if int(userdata['headless'])==1 :
  driver.set_window_position(-10000,0)
 user,password,rs=userdata['usereway'],userdata['passwordeway'],userdata['rs']
 #website=userdata['website']
 website=''
 #driver.set_window_position(-10000,0)
 driver.get('https://leveredge102.hulcd.com/rsunify/')
 searchbox = driver.find_element_by_xpath('//*[@id="userName"]')
 searchbox.send_keys('CREDIT')
 searchbox1 = driver.find_element_by_xpath('//*[@id="password"]')
 searchbox1.send_keys('Ven2004@')
 searchbox = driver.find_element_by_xpath('//*[@id="databaseName"]')
 searchbox.send_keys(rs)
 but = driver.find_element_by_xpath('//*[@id="gologin"]')
 but.click()
 t2=0
 while True:
   t2+=1
   try :
      captcha=driver.find_element(By.ID,'cap_question')
      captcha=captcha.get_attribute('innerText')
      captcha=captcha.replace('=','')
      captcha=captcha.split('+')
      try :
       captcha=[int(i.strip()) for i in captcha]
      except :
        time.sleep(0.5)
        continue
      value=sum(captcha)
      driver.execute_script('document.getElementById("cap_answer").value='+str(value))
      driver.execute_script('confirmSubmission();')
      print(1)
      break 
   except Exception as e1:
        #print(e1)
        for t1 in range(0,1) :
          try :
            driver.execute_script('document.querySelector("#ikea_home_menu_search").click();')
          except : 
           time.sleep(1)
        if t1 == 0 :
            continue 
        break
 time.sleep(0.5)
 date1 = datetime.datetime.now().strftime('%d/%m/%Y')
 date2=date1.split('/')[2]+'-'+date1.split('/')[1]+'-'+date1.split('/')[0]
 intial=os.listdir(path)
 f=open(r'D:\currentstock.txt')
 x=f.read()
 f.close()
 x=x.replace('_from1_',date1)
 x=x.replace('_from2_',date2)
 driver.execute_script(x)
 while True :
   if len(os.listdir(path))== len(intial)+1 :
     try :
         file1=list((set(os.listdir(path))^set(intial))&set(os.listdir(path)))[0]
         if 'tmp' not in file1 and 'crd' not in file1:
           df = pd.read_excel(path+file1)
           #print(df)
           break
         else :
             time.sleep(0.5)
     except Exception as e:
        #print(e)
        continue
   else :
      time.sleep(0.5) 
 time.sleep(0.2)
 driver.find_element(By.ID,"ikea_home_menu_search").send_keys("basepack")
 waituntil('document.querySelector("#search_menu_item_container > div:nth-child(1)").click();')
 driver.switch_to.window(driver.window_handles[1])
 waituntil('exportData_Basepack()')
 intial=os.listdir(path)
 while True :
   if len(os.listdir(path))== len(intial)+1 :
     try :
         file1=list((set(os.listdir(path))^set(intial))&set(os.listdir(path)))[0]
         if 'tmp' not in file1 and 'crd' not in file1:
           #df1 = pd.read_excel(path+file1,)
           wb = load_workbook(path+file1,data_only=True)
           #print(df1)
           break
         else :
             time.sleep(0.5)
     except Exception as e:
        print(e)
        continue
   else :
      time.sleep(0.5)
 def conv(x) :
    try :
     return str(int(x))
    except :
     return '' 
 #df1['BasePack Code']=df1['BasePack Code'].apply(lambda x: conv(x))
 data = {}
 def change(code,status,mini) :
  df2=df.copy()
  df2 = df2[df2[df2.columns[1]]==code]
  #print(code)
  df2=df2.reset_index(drop=True)
  if len(df2.index)==0 :
      real = 'INACTIVE'
  else :
    if float(df2.iloc[0][df.columns[13]]) >= float(mini) :
        #if code=='88139' :
          #print(df2)
          #print(float(df2.iloc[0][df.columns[13]]) , float(mini))
        real='ACTIVE'
    else :
        real='INACTIVE'
  return real
 final=[]
 ws = wb.worksheets[0]
 i=0
 for row in ws.iter_rows() :
  if i==0 :
    i=1
    continue
  i+=1
  #print(row[5].fill.start_color.index,type(row[5].fill.start_color.index))
  if row[1].value==None :
      continue 
  if row[5].fill.start_color.index != '52' :    
   code,status=row[5].value,row[10].value
   #print(code,1)
   if code=='' :
    continue
   if code in data.keys():
     row[10].value=change(code,status,data[code]) 
   else :
     row[10].value=change(code,status,1)
 ws.delete_cols(1, 5)
 ws.delete_cols(7,3)
 wb.save(filename = path+'basepack_updated.xlsx')
 excel=dispatch('Excel.Application')
 excel.DisplayAlerts = 0
 wb = excel.Workbooks.Open(path+'basepack_updated.xlsx')
 wb.SaveAs(path+'basepack_updated.xlsx')
 wb.Close()
 excel.Application.Quit()
 driver.find_element(By.ID,'attach_file').send_keys(path+'basepack_updated.xlsx')
 driver.quit()

 
